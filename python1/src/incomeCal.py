# 基于持仓表每日预期收益 × 持仓天数
import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime

# 加载Excel工作簿
file_path = r"C:\Users\22938\Documents\理财.xlsx"  # 替换为您的文件路径
wb = openpyxl.load_workbook(file_path, data_only=True)

# 获取持仓数据 (Holdings sheet)
holdings_sheet = wb["Holdings"]
holdings_data = []

for row in range(2, holdings_sheet.max_row + 1):
    trade_date = holdings_sheet.cell(row, column_index_from_string("E")).value  # 交易日期
    end_date = holdings_sheet.cell(row, column_index_from_string("G")).value    # 到期日期
    daily_return = holdings_sheet.cell(row, column_index_from_string("H")).value  # 日收益
    
    if all([trade_date, daily_return]):  # 确保必要数据存在
        holdings_data.append({
            "trade_date": trade_date,
            "end_date": end_date if end_date else datetime.today().date(),
            "daily_return": daily_return
        })

# 计算Income sheet D列的值
income_sheet = wb["Income"]
today = datetime.today().date()

for row in range(2, income_sheet.max_row + 1):
    month_start = income_sheet.cell(row, 2).value  # B列 - 月初
    month_end = income_sheet.cell(row, 3).value    # C列 - 月末
    total_income = 0.0
    
    # 计算当月所有产品的收益
    for product in holdings_data:
        # 计算持仓开始日期（取较大值）
        start_date = max(
            month_start, 
            product["trade_date"]
        )
        
        # 计算持仓结束日期（取较小值）
        end_date = min(
            month_end, 
            product["end_date"]
        )
        
        # 计算持仓天数（忽略负值）
        holding_days = (end_date - start_date).days + 1
        if holding_days < 0:
            holding_days = 0
        
        # 累加收益
        total_income += product["daily_return"] * holding_days
    
    # 更新D列值
    income_sheet.cell(row, 4).value = total_income  # D列

# 保存计算结果
wb.save(file_path)
print(f"处理完成！已更新{income_sheet.max_row-1}行数据。")