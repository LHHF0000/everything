import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import List, Tuple, Optional, Dict, Any
import re


class ExcelToMarkdownConverter:
    def __init__(self, file_path: str):
        """
        初始化转换器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.merged_cells_info = {}  # 存储合并单元格信息
        self.cell_value_cache = {}  # 缓存单元格值（处理合并单元格）
    
    def load_workbook(self, sheet_name: Optional[str] = None) -> None:
        """
        加载工作簿
        
        Args:
            sheet_name: 工作表名称，如果为None则使用活动工作表
        """
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[sheet_name]
            else:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
        else:
            self.worksheet = self.workbook.active
    
    def get_merged_cells_info(self) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
        """
        获取所有合并单元格的信息
        
        Returns:
            字典，键为单元格坐标(row, col)，值为合并范围(min_row, min_col, max_row, max_col)
        """
        merged_info = {}
        
        for merged_range in self.worksheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = (
                merged_range.min_row, 
                merged_range.min_col, 
                merged_range.max_row, 
                merged_range.max_col
            )
            
            # 获取合并单元格左上角的值
            main_cell_value = self.worksheet.cell(row=min_row, column=min_col).value
            
            # 存储合并范围信息
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_info[(row, col)] = (min_row, min_col, max_row, max_col, main_cell_value)
        
        return merged_info
    
    def get_cell_value(self, row: int, col: int) -> Any:
        """
        获取单元格值，处理合并单元格
        
        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）
            
        Returns:
            单元格值
        """
        # 如果单元格在合并范围内，返回合并单元格左上角的值
        if (row, col) in self.merged_cells_info:
            _, _, _, _, main_value = self.merged_cells_info[(row, col)]
            return main_value
        
        # 否则返回实际单元格值
        return self.worksheet.cell(row=row, column=col).value
    
    def parse_range_string(self, range_str: str) -> Tuple[int, int, int, int]:
        """
        解析Excel范围字符串（如"A1:D10"）
        
        Args:
            range_str: Excel范围字符串
            
        Returns:
            (start_row, start_col, end_row, end_col)
        """
        # 匹配类似"A1:D10"的格式
        pattern = r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$'
        match = re.match(pattern, range_str.upper())
        
        if not match:
            raise ValueError(f"无效的范围格式: {range_str}")
        
        start_col_str, start_row_str, end_col_str, end_row_str = match.groups()
        
        start_col = column_index_from_string(start_col_str)
        end_col = column_index_from_string(end_col_str)
        start_row = int(start_row_str)
        end_row = int(end_row_str)
        
        return start_row, start_col, end_row, end_col
    
    def build_tree_structure(self, 
                           start_row: int, 
                           start_col: int, 
                           end_row: int, 
                           end_col: int) -> List[Dict[str, Any]]:
        """
        构建树状结构
        
        Args:
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
            
        Returns:
            树状结构数据
        """
        tree_data = []
        
        for row in range(start_row, end_row + 1):
            # 当前行的数据
            row_data = {}
            
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                level = col - start_col  # 层级（0为第一级，1为第二级，以此类推）
                
                # 获取单元格值
                cell_value = self.get_cell_value(row, col)
                
                # 存储到行数据中
                row_data[level] = {
                    'col': col,
                    'col_letter': col_letter,
                    'row': row,
                    'value': cell_value if cell_value is not None else '',
                    'level': level
                }
            
            # 添加到数据列表
            tree_data.append(row_data)
        
        return tree_data
    
    def convert_to_markdown(self, 
                          tree_data: List[Dict[str, Any]], 
                          max_level: int) -> str:
        """
        将树状结构转换为Markdown格式
        
        Args:
            tree_data: 树状结构数据
            max_level: 最大层级（列数）
            
        Returns:
            Markdown格式字符串
        """
        markdown_lines = []
        last_values = {}  # 记录上一行各列的值，用于判断是否重复
        
        for i, row_data in enumerate(tree_data):
            current_line = ""
            previous_value_changed = False  # 标记上一级的值是否变化
            
            for level in range(max_level + 1):
                if level in row_data:
                    cell_info = row_data[level]
                    value = cell_info['value']
                    
                    # 如果值为空，跳过
                    if not value:
                        continue
                    
                    # 检查当前值是否与上一行相同
                    is_same_as_previous = (
                        i > 0 and 
                        level in last_values and 
                        last_values[level] == value
                    )
                    
                    # 如果当前层级的值与上一行相同，且上一级的值没有变化，则跳过
                    if is_same_as_previous and not previous_value_changed:
                        continue
                    
                    # 如果当前层级的值与上一行不同，标记上一级的值已变化
                    if not is_same_as_previous:
                        previous_value_changed = True
                    
                    # 添加缩进（每个层级缩进2个空格）
                    indent = "#" * level
                    
                    # 如果是第一级，使用一级标题
                    if level == 0:
                        current_line += f"{indent}# {value}\n"
                    # 其他级别，使用列表项
                    else:
                        current_line += f"{indent}* {value}\n"
                    
                    # 更新上一行的值记录
                    last_values[level] = value
            
            # 将当前行添加到Markdown结果
            if current_line:
                markdown_lines.append(current_line)
        
        return "".join(markdown_lines)
    
    def convert(self, 
               range_str: Optional[str] = None,
               sheet_name: Optional[str] = None) -> str:
        """
        主转换方法
        
        Args:
            range_str: Excel范围字符串（如"A1:D10"），如果为None则读取整个工作表
            sheet_name: 工作表名称，如果为None则使用活动工作表
            
        Returns:
            Markdown格式字符串
        """
        # 加载工作簿
        self.load_workbook(sheet_name)
        
        # 获取合并单元格信息
        self.merged_cells_info = self.get_merged_cells_info()
        
        # 确定读取范围
        if range_str:
            start_row, start_col, end_row, end_col = self.parse_range_string(range_str)
        else:
            # 如果没有指定范围，读取整个工作表的有数据区域
            start_row = 1
            start_col = 1
            end_row = self.worksheet.max_row
            end_col = self.worksheet.max_column
        
        print(f"读取范围: {get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
        print(f"行数: {end_row - start_row + 1}, 列数: {end_col - start_col + 1}")
        
        # 构建树状结构
        tree_data = self.build_tree_structure(start_row, start_col, end_row, end_col)
        
        # 计算最大层级
        max_level = end_col - start_col
        
        # 转换为Markdown
        markdown = self.convert_to_markdown(tree_data, max_level)
        
        return markdown
    
    def save_markdown(self, markdown_text: str, output_path: str) -> None:
        """
        保存Markdown到文件
        
        Args:
            markdown_text: Markdown文本
            output_path: 输出文件路径
        """
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown_text)
        
        print(f"Markdown已保存到: {output_path}")


def main():
    """
    使用示例
    """
    # Excel文件路径
    excel_file = "C:\\Users\\22938\\Downloads\\评分规则 - 标注 - 删减版\\附件：客维业务服务品质评价标准.xlsx"
    
    # 创建转换器
    converter = ExcelToMarkdownConverter(excel_file)
    
    try:
        # 示例2: 转换指定范围
        print("\n" + "=" * 50)
        print("示例2: 转换指定范围 B4:F29")
        print("=" * 50)
        markdown2 = converter.convert(range_str="B4:F29")
        print(markdown2)
        converter.save_markdown(markdown2, "output_range.md")
        
        
    except Exception as e:
        print(f"转换过程中发生错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()