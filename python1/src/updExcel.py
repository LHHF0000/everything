import openpyxl
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def deal(s):
    i = 2
    while i < rows + 1:
        for row_b in worksheet_b.iter_rows(min_row=2, values_only=True):
            table_name_b = row_b[0]
            column_name_b = row_b[1]
            data_type_b = row_b[2]

            col1 = sheet1.cell(row=i, column=s-2).value  # 获取(row,col)单元格的值
            col2 = sheet1.cell(row=i, column=s-1).value
            col3 = sheet1.cell(row=i, column=s).value
            if col1 != table_name_b or col2 != column_name_b:
                continue
            if col1 == table_name_b and col2 == column_name_b and col3 != data_type_b:
                sheet1.cell(row=i, column=s, value=data_type_b).fill = filleN
                break
            if col1 == table_name_b and col2 == column_name_b and col3 == data_type_b:
                sheet1.cell(row=i, column=s, value=data_type_b).fill = filleY
                break
            # ws.cell(row=i, column=7).value = 'varchar(2)'
        # ws.cell(row=row, column=col).value =
        i += 1


if __name__ == '__main__':
    excelA_path = r'D:\everything\python1\files\excelA.xlsx'
    excelB_path = r'D:\everything\python1\files\excelB.xlsx'
    excelA_updated_path = r'D:\everything\python1\files\excelA_updated.xlsx'

    # 复制excelA.xlsx为excelA_updated.xlsx
    shutil.copyfile(excelA_path, excelA_updated_path)
    workbook_b = load_workbook(excelB_path)
    worksheet_b = workbook_b.active

    wb = openpyxl.load_workbook(excelA_path)
    sheet1 = wb.worksheets[0]  # 获取第一页sheet

    cols = sheet1.max_column  # 获取有效的数据的最大列数
    rows = sheet1.max_row  # 最大行数

    filleY = PatternFill('solid', fgColor='a5b46a')  # 黄绿
    filleN = PatternFill('solid', fgColor='6ab479')  # 绿
    deal(7)
    deal(11)
    deal(15)

        # 此处有一个注意事项
        # cell(r,c)中的行r和列c 是从1开始。而不是0。  例如A11则是ws.cell(1,1),不是ws.cell(0,0)
    # wb = openpyxl.load_workbook(xlsx_path)
    wb.save(excelA_updated_path)  # 设置保存的路径 和保存的文件名：2.xlsx